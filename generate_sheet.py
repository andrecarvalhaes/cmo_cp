import json, base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

with open('/tmp/designer_final.json', encoding='utf-8') as f:
    final_results = json.load(f)

MONTHS_ORDER = ['Janeiro', 'Fevereiro', 'Março', 'Abril']
DESIGNERS = ['Felipe Galo', 'Ederson Carlos', 'Henrique Felix']

wb = Workbook()

# Styles
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2D3436', end_color='2D3436', fill_type='solid')
light_gray = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD'),
)

# ========== ABA RESUMO ==========
ws_resumo = wb.active
ws_resumo.title = 'Resumo'

# Title
ws_resumo.merge_cells('A1:F1')
ws_resumo['A1'] = 'Kanban Design - % Entregas no Prazo (Jan-Abr 2026)'
ws_resumo['A1'].font = Font(bold=True, size=14)
ws_resumo['A1'].alignment = Alignment(horizontal='center')

# Headers
headers_resumo = ['Designer', 'Janeiro', 'Fevereiro', 'Marco', 'Abril', 'Geral']
for col, h in enumerate(headers_resumo, 1):
    cell = ws_resumo.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data
for row_idx, designer in enumerate(DESIGNERS, 4):
    rows = final_results[designer]
    ws_resumo.cell(row=row_idx, column=1, value=designer).font = Font(bold=True, size=11)
    ws_resumo.cell(row=row_idx, column=1).border = thin_border

    total_on_time = 0
    total_valid = 0

    for col_idx, month in enumerate(MONTHS_ORDER, 2):
        month_rows = [r for r in rows if r['mes'] == month]
        valid = [r for r in month_rows if r['is_late'] is not None]
        on_time = sum(1 for r in valid if not r['is_late'])
        total = len(valid)

        total_on_time += on_time
        total_valid += total

        cell = ws_resumo.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        if total > 0:
            pct = on_time / total * 100
            cell.value = f'{pct:.0f}% ({on_time}/{total})'
            if pct >= 95:
                cell.font = Font(bold=True, color='00B894')
            elif pct >= 85:
                cell.font = Font(bold=True, color='F39C12')
            else:
                cell.font = Font(bold=True, color='D63031')
        else:
            cell.value = 'N/A'

    # Geral
    cell = ws_resumo.cell(row=row_idx, column=6)
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
    if total_valid > 0:
        pct = total_on_time / total_valid * 100
        cell.value = f'{pct:.0f}% ({total_on_time}/{total_valid})'
        cell.font = Font(bold=True, size=12)

# Detail table
row = 8
ws_resumo.merge_cells(f'A{row}:F{row}')
ws_resumo.cell(row=row, column=1, value='Detalhamento').font = Font(bold=True, size=12)
row += 1

sub_headers = ['Designer', 'Mes', 'Total', 'No Prazo', 'Atrasadas', '% No Prazo']
for col, h in enumerate(sub_headers, 1):
    cell = ws_resumo.cell(row=row, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border
row += 1

for designer in DESIGNERS:
    rows = final_results[designer]
    for month in MONTHS_ORDER:
        month_rows = [r for r in rows if r['mes'] == month]
        valid = [r for r in month_rows if r['is_late'] is not None]
        on_time = sum(1 for r in valid if not r['is_late'])
        late = sum(1 for r in valid if r['is_late'])
        total = len(valid)
        pct = f'{on_time/total*100:.0f}%' if total else 'N/A'

        ws_resumo.cell(row=row, column=1, value=designer).border = thin_border
        ws_resumo.cell(row=row, column=2, value=month).border = thin_border
        ws_resumo.cell(row=row, column=3, value=total).border = thin_border
        ws_resumo.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row, column=4, value=on_time).border = thin_border
        ws_resumo.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row, column=5, value=late).border = thin_border
        ws_resumo.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws_resumo.cell(row=row, column=6, value=pct).border = thin_border
        ws_resumo.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        row += 1

# Column widths
ws_resumo.column_dimensions['A'].width = 20
for c in ['B', 'C', 'D', 'E', 'F']:
    ws_resumo.column_dimensions[c].width = 18

# ========== ABAS INDIVIDUAIS ==========
detail_headers = ['Mes', 'Tarefa', 'Due Date', 'Entrega (Checar)', 'Status', 'Link']

for designer in DESIGNERS:
    ws = wb.create_sheet(title=designer)
    rows = final_results[designer]

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'{designer} - Tarefas Concluidas Jan-Abr 2026'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Headers
    for col, h in enumerate(detail_headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for i, r in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=r['mes']).border = thin_border
        ws.cell(row=i, column=2, value=r['name']).border = thin_border
        ws.cell(row=i, column=3, value=r['due']).border = thin_border
        ws.cell(row=i, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=i, column=4, value=r['checar']).border = thin_border
        ws.cell(row=i, column=4).alignment = Alignment(horizontal='center')

        cell = ws.cell(row=i, column=5, value=r['status'])
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
        if r['status'] == 'No prazo':
            cell.font = Font(color='00B894', bold=True)
        elif r['status'] == 'Atrasada':
            cell.font = Font(color='D63031', bold=True)

        cell = ws.cell(row=i, column=6)
        cell.value = r['link']
        cell.hyperlink = r['link']
        cell.font = Font(color='0984E3', underline='single')
        cell.border = thin_border

    # Alternating rows
    for i in range(4, 4 + len(rows)):
        if i % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=i, column=col).fill = light_gray

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 40

    print(f'{designer}: {len(rows)} rows')

# Save
output_path = '/tmp/kanban_design_2026.xlsx'
wb.save(output_path)
print(f'\nSaved to {output_path}')

# Base64 for Google Drive upload
with open(output_path, 'rb') as f:
    b64 = base64.b64encode(f.read()).decode()
with open('/tmp/kanban_b64.txt', 'w') as f:
    f.write(b64)
print(f'Base64 ready ({len(b64)} chars)')
